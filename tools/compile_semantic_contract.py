#!/usr/bin/env python3
"""Compile the semantic contract, and refuse to if the artifacts disagree.

**What this is.** `semantic/ontology/terms.xlsx` is where the grammar, seed
vocabulary, aliases, fixtures and runtime configuration are *authored*.
`semantic/contracts/mention_extract_v1.schema.json` is the strict model-output
contract. Neither is what production reads. Production reads
`semantic/contracts/compiled_semantic_contract_v1.json`, and this file is the
only thing allowed to write it.

**Why a compiler rather than three files kept in step by hand.** This repository
has paid for the alternative three times. `SOURCE_ACTION_PAIRS` in Python and
`sources.action_weights` in SQL disagreed about `top_track` for months, and the
comment recording that defect sat directly above the identical bug for
`playlist_item`. `0191` published a genre mint and revoked every privilege on it,
so the migration's own header described behaviour the shipped code could not
perform. **One fact in two places is the recurring defect**, and the response is
to derive rather than duplicate.

**Determinism is the load-bearing property.** The deploy validator recompiles the
contract and compares canonical content after removing only `generated_at`. That
comparison is worthless unless the same inputs always produce the same bytes, so
every collection here is emitted in a defined order — workbook order where the
order carries meaning, sorted where it does not — and nothing reads the clock
except the one field that is excluded from the comparison.

    python3 tools/compile_semantic_contract.py --check        # verify, emit nothing
    python3 tools/compile_semantic_contract.py --emit         # rewrite the contract
    python3 tools/compile_semantic_contract.py --check-database --live-enums live.json

The database gate is separate and takes its live values as an argument; see
`check_database` for why it does not open a connection of its own.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import pathlib
import sys
from typing import Any

REPOSITORY = pathlib.Path(__file__).resolve().parent.parent
WORKBOOK = REPOSITORY / "semantic" / "ontology" / "terms.xlsx"
#: **The schema is chosen by the workbook, not by this constant.** It was
#: `mention_extract_v1.schema.json` here, so the only way to evaluate a new
#: output contract was to edit the compiler — and `llm.output.schema_version`
#: already said which schema was authoritative, one fact in two places with the
#: code winning. `output_schema_path` derives the file from that key, so naming a
#: schema and compiling against it are the same act.
SCHEMA_DIR = REPOSITORY / "semantic" / "contracts"
CONTRACT = REPOSITORY / "semantic" / "contracts" / "compiled_semantic_contract_v1.json"

CONTRACT_VERSION = "compiled_semantic_contract_v1"

# **Families that describe evidence rather than vocabulary.** A video, an
# episode or a calendar event is a thing we observed, not a term anybody could
# be interested in, so none of them may map to a concept row. They are listed
# here because the coverage check demands that *every* family resolve exactly
# once, and "resolves to nothing, on purpose" is an answer that has to be
# written down to be distinguishable from an omission.
VIRTUAL_FAMILIES = frozenset({
    "video", "episode", "article", "observation",
    "calendar_event", "travel_itinerary", "event_occurrence", "user_profile",
})

# The ontology families the model may never emit. Checked as an exact difference
# rather than a subset test: a family appearing on one side and not the other is
# a drift signal whichever direction it points.
#
# **`music_recording` is the sixth, and it is a different kind of entry from the
# other five.** Those are plumbing the model has no business proposing — a
# channel is provider identity, a hub is navigation. This one is a decision:
# `0221` took recordings out of the versioned ontology because identity is not
# vocabulary, and `mention_extract_v2` dropped the family so a model could not
# reopen the minting route that closed. The ontology still holds it, for the
# identity registry and for `provisional_entities`; what ended is the model's
# licence to propose one.
MODEL_FORBIDDEN_FAMILIES = frozenset({
    "channel", "event_type", "game_category", "hub", "platform",
    "music_recording",
})


class ContractError(RuntimeError):
    """A refusal to compile. Never repaired automatically."""


def _sha256(path: pathlib.Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _split(value: str, separator: str = "|") -> list[str]:
    return [part.strip() for part in str(value).split(separator) if part.strip()]


def load_workbook() -> dict[str, Any]:
    """Read the authoring workbook into plain Python.

    `openpyxl` is a dev dependency and deliberately not a runtime one: the
    worker consumes the compiled JSON and never opens a spreadsheet.
    """
    try:
        import openpyxl
    except ModuleNotFoundError as missing:  # pragma: no cover - environment
        raise ContractError(
            "openpyxl is required to compile the contract; it is in the dev extra"
        ) from missing

    book = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    sheets: dict[str, list[list[str]]] = {}
    for name in ("terms", "relations", "aliases", "grammar", "examples", "runtime_config"):
        if name not in book.sheetnames:
            raise ContractError(f"workbook is missing the {name!r} sheet")
        rows = [
            ["" if cell is None else str(cell).strip() for cell in row]
            for row in book[name].iter_rows(values_only=True)
        ]
        sheets[name] = [row for row in rows if any(row)]
    return sheets


def config_of(sheets: dict[str, Any]) -> dict[str, str]:
    rows = sheets["runtime_config"]
    return {row[0]: (row[1] if len(row) > 1 else "") for row in rows[1:] if row and row[0]}


def _column(sheet: list[list[str]], name: str) -> int:
    header = sheet[0]
    if name not in header:
        raise ContractError(f"expected a {name!r} column, found {header}")
    return header.index(name)


def require(config: dict[str, str], key: str) -> str:
    """A missing key is a compile failure, never a default.

    A default here would be the third instance of the defect this file exists to
    prevent: the compiler would quietly supply a value the workbook does not
    state, and the two would drift with nothing reporting it.
    """
    if key not in config or config[key] == "":
        raise ContractError(f"runtime_config is missing {key!r}")
    return config[key]


def request_schema_path(config: dict[str, str]) -> pathlib.Path:
    """The request schema the workbook says is authoritative.

    Its counterpart. `llm.input.schema_version` named the *response* schema
    until `0241`'s gateway work, which was not wrong so much as vacant — nothing
    read it and there was no request schema for it to name.
    """
    version = require(config, "llm.input.schema_version")
    path = SCHEMA_DIR / f"{version}.schema.json"
    if not path.exists():
        raise ContractError(
            f"llm.input.schema_version names {version!r} and "
            f"{path.name} does not exist"
        )
    return path


def output_schema_path(config: dict[str, str]) -> pathlib.Path:
    """The response schema the workbook says is authoritative.

    A missing file is a `ContractError` rather than an `OSError`, because the
    likely cause is a version named before its schema was written and the
    author should be told which name failed.
    """
    version = require(config, "llm.output.schema_version")
    path = SCHEMA_DIR / f"{version}.schema.json"
    if not path.exists():
        raise ContractError(
            f"llm.output.schema_version names {version!r} and "
            f"{path.name} does not exist"
        )
    return path


def model_predicates(sheets: dict[str, Any]) -> list[str]:
    """Predicates the grammar permits the model to propose, in grammar order.

    Order is the sheet's, not sorted, because `ontology_compiler.model_predicates`
    records *what the grammar says* and the grammar is an authored document whose
    row order a reviewer reads.
    """
    grammar = sheets["grammar"]
    predicate = _column(grammar, "predicate")
    propose = _column(grammar, "llm_may_propose")
    seen: list[str] = []
    for row in grammar[1:]:
        if len(row) > propose and row[propose].upper() == "T":
            key = row[predicate]
            if key and key not in seen:
                seen.append(key)
    return seen


def _mention_properties(schema: dict[str, Any]) -> dict[str, Any]:
    """The mention property set, under either schema encoding.

    The tags/index conditional became two anyOf variants (`mention_text`,
    `mention_tag`) because xgrammar cannot compile `if/then` and its matcher
    leaks on patterned strings (2026-08-19). The variants must agree on every
    enum this compiler checks — a value present in one door and not the other
    would let the same response be legal or illegal depending on which field it
    cites — so disagreement is refused here rather than discovered at parity.
    """
    defs = schema["$defs"]
    if "mention" in defs:
        return defs["mention"]["properties"]

    # **Every variant present, not the two this function was written for.**
    # v3 adds `mention_inferred`, and a third door that nothing compared would
    # be exactly the hole this check exists to close: a family legal when
    # inferred and illegal when read, discovered by a refusal in production.
    # Named from the schema rather than listed here, so a fourth variant is
    # compared the day it appears.
    variants = {name: body["properties"]
                for name, body in defs.items()
                if name.startswith("mention_")}
    if not variants:
        raise ContractError("the schema declares no mention variant")

    first_name, first = next(iter(sorted(variants.items())))
    for name, properties in sorted(variants.items()):
        for key in ("family_hypothesis", "mention_role"):
            if properties[key]["enum"] != first[key]["enum"]:
                raise ContractError(
                    f"mention variants disagree on {key}: "
                    f"{name} against {first_name}")
    return first


def _abstain_reasons(schema: dict[str, Any]) -> list[str]:
    """The non-null abstain vocabulary, under either schema encoding."""
    defs = schema["$defs"]
    item = defs.get("item_abstained") or defs["item"]
    return [v for v in item["properties"]["abstain_reason"]["enum"] if v]


def validate(sheets: dict[str, Any], schema: dict[str, Any],
             request_schema: dict[str, Any] | None = None) -> None:
    """Every check that must hold before a contract may be written.

    Each raises rather than warns. A compiler that emits a contract while
    reporting a problem is a compiler whose output nobody reads.
    """
    config = config_of(sheets)
    mention = _mention_properties(schema)
    schema_enums = {
        "llm.family.enum": mention["family_hypothesis"]["enum"],
        "llm.mention_role.enum": mention["mention_role"]["enum"],
        "llm.schema.abstain_reasons": _abstain_reasons(schema),
    }

    # 1. Enum parity, both directions. Sorted equality rather than a subset test:
    #    a value in the workbook and not the schema is as wrong as the reverse,
    #    and only one of those is visible from the model's side.
    # **A relation vocabulary is not an extraction fact.** `mention_extract_v2`
    # emits no `relation_hypothesis` at all, so reading the predicate list off
    # the schema is reading a `$def` that is not there. The grammar sheet is the
    # authority either way — `ontology_compiler.model_predicates` has always been
    # sourced from it — and the schema is asked only when it actually declares
    # what the model may propose.
    relations = schema["$defs"].get("relation_hypothesis")
    if relations is not None:
        schema_enums["llm.predicate.enum"] = relations["properties"]["predicate"]["enum"]

    for key, values in schema_enums.items():
        workbook_values = sorted(_split(require(config, key)))
        if workbook_values != sorted(values):
            only_workbook = sorted(set(workbook_values) - set(values))
            only_schema = sorted(set(values) - set(workbook_values))
            raise ContractError(
                f"{key} disagrees with the JSON Schema; "
                f"workbook-only={only_workbook} schema-only={only_schema}"
            )

    # 1b. **The request schema's bounds are the workbook's.** Every number in it
    #     is authored under `llm.input.*` and `llm.batch.*`, and a schema that
    #     drifted from them would bound the wire at one size while the prompt and
    #     the budget report assumed another.
    if request_schema is not None:
        fields = request_schema["$defs"]["fields"]["properties"]
        expected_bounds = {
            "title": int(require(config, "llm.input.max_title_chars")),
            "channel_label": int(require(config, "llm.input.max_channel_chars")),
            "description_excerpt":
                int(require(config, "llm.input.max_description_chars")),
            # The music lane's stated context (2026-08-21): what Apple says
            # about the row, so a song can nominate its film or artist.
            "performer": int(require(config, "llm.input.max_performer_chars")),
            "composer": int(require(config, "llm.input.max_composer_chars")),
            "album": int(require(config, "llm.input.max_album_chars")),
        }
        for field, bound in expected_bounds.items():
            if fields[field]["maxLength"] != bound:
                raise ContractError(
                    f"request schema bounds {field} at "
                    f"{fields[field]['maxLength']} and the workbook says {bound}"
                )
        if fields["tags"]["maxItems"] != int(require(config, "llm.input.max_tags")):
            raise ContractError("request schema and workbook disagree on max_tags")
        if fields["tags"]["items"]["maxLength"] != int(
                require(config, "llm.input.max_tag_chars")):
            raise ContractError("request schema and workbook disagree on max_tag_chars")
        wire = int(require(config, "llm.batch.max_items"))
        if request_schema["properties"]["items"]["maxItems"] != wire:
            raise ContractError(
                "the request schema admits a different number of items than the wire maximum")
        if schema["properties"]["items"]["maxItems"] != wire:
            raise ContractError(
                "the response schema admits a different number of items than the wire maximum")

    # 2. The grammar's propose flags are the same twelve predicates. This is the
    #    check that would have caught the prompt offering seven predicates while
    #    the schema allowed twelve, which silently made every sports roster
    #    relation unreachable.
    grammar_predicates = model_predicates(sheets)
    if "llm.predicate.enum" in schema_enums:
        if sorted(grammar_predicates) != sorted(schema_enums["llm.predicate.enum"]):
            raise ContractError(
                "grammar llm_may_propose=T predicates disagree with the schema: "
                f"grammar={sorted(grammar_predicates)} "
                f"schema={sorted(schema_enums['llm.predicate.enum'])}"
            )
    elif sorted(grammar_predicates) != sorted(_split(require(config, "llm.predicate.enum"))):
        raise ContractError(
            "grammar llm_may_propose=T predicates disagree with the workbook: "
            f"grammar={sorted(grammar_predicates)} "
            f"workbook={sorted(_split(require(config, 'llm.predicate.enum')))}"
        )

    # 3. A source profile may narrow the union; it may never widen it.
    union = set(grammar_predicates)
    for key, value in config.items():
        if key.startswith("llm.predicate.profile."):
            extra = sorted(set(_split(value)) - union)
            if extra:
                raise ContractError(f"{key} proposes predicates outside the union: {extra}")

    # 4. Every family resolves exactly once, and virtual families resolve to
    #    nothing on purpose.
    mappings = {
        key[len("term_family.map."):]: value
        for key, value in config.items()
        if key.startswith("term_family.map.") and key != "term_family.map.version"
    }
    declared = set(_split(require(config, "ontology.family.enum")))
    term_families = {
        row[_column(sheets["terms"], "family")]
        for row in sheets["terms"][1:]
        if len(row) > _column(sheets["terms"], "family") and row[_column(sheets["terms"], "family")]
    }
    for family in sorted(term_families | declared):
        if family in VIRTUAL_FAMILIES:
            if family in mappings:
                raise ContractError(f"virtual family {family!r} must not map to storage")
            continue
        if family not in mappings:
            raise ContractError(f"family {family!r} has no term_family.map entry")
    unknown = sorted(set(mappings) - declared - VIRTUAL_FAMILIES)
    if unknown:
        raise ContractError(f"term_family.map covers families absent from the ontology: {unknown}")

    # 5. What the model may emit is a strict subset of what the ontology holds,
    #    and the difference is exactly the five it must never produce.
    model_families = set(schema_enums["llm.family.enum"])
    difference = declared - model_families
    if difference != set(MODEL_FORBIDDEN_FAMILIES):
        raise ContractError(
            "the families the model may not emit are not the expected five: "
            f"{sorted(difference)}"
        )
    if model_families - declared:
        raise ContractError(
            f"the model may emit families the ontology does not hold: "
            f"{sorted(model_families - declared)}"
        )

    # 6. Hub canonicalization: nothing legacy survives, and every canonical
    #    target is a production hub id rather than an authoring one.
    canonical = {
        key[len("ontology.hub.canonical."):]: value
        for key, value in config.items()
        if key.startswith("ontology.hub.canonical.")
    }
    redirects = {
        key[len("ontology.hub.alias."):]: value
        for key, value in config.items()
        if key.startswith("ontology.hub.alias.")
        and not key.endswith((".normalization", ".prompt_policy"))
    }
    for source, target in sorted(redirects.items()):
        if target not in set(canonical.values()):
            raise ContractError(f"hub redirect {source!r} points outside the canonical set")
        if target in redirects:
            raise ContractError(f"hub redirect {source!r} points at another redirect")
    hub_column = _column(sheets["terms"], "subject_ID")
    family_column = _column(sheets["terms"], "family")
    for row in sheets["terms"][1:]:
        if len(row) > family_column and row[family_column] == "hub":
            key = row[hub_column]
            if key in redirects:
                raise ContractError(
                    f"the workbook still authors the legacy hub {key!r}; "
                    f"use {redirects[key]!r}"
                )

    # 7. The packing formula must be arithmetically consistent with the values
    #    beside it, or the startup assertion it describes is decorative.
    reserve = int(require(config, "llm.output.uncalibrated_item_reserve_tokens"))
    envelope = int(require(config, "llm.output.envelope_token_reserve"))
    ceiling = int(require(config, "llm.max_output_tokens"))
    headroom = float(require(config, "llm.output.packing_headroom"))
    calibrated = int(require(config, "llm.batch.calibrated_max_items"))
    derived = int((ceiling - envelope) // (reserve * (1 + headroom)))
    if calibrated > derived:
        raise ContractError(
            f"calibrated_max_items={calibrated} exceeds the {derived} items the "
            f"budget permits ({ceiling}-{envelope})/({reserve}*{1 + headroom})"
        )
    if int(require(config, "llm.batch.default_items")) > calibrated:
        raise ContractError("default_batch_items exceeds calibrated_max_items")


def compile_contract(sheets: dict[str, Any], schema: dict[str, Any],
                     schema_path: pathlib.Path, generated_at: str) -> dict[str, Any]:
    config = config_of(sheets)
    mention = _mention_properties(schema)

    return {
        "contract_version": CONTRACT_VERSION,
        "generated_at": generated_at,
        "source_hashes": {
            # Recomputed, never copied. A hash carried in the workbook would be a
            # claim about the workbook made by the workbook.
            "workbook_sha256": _sha256(WORKBOOK),
            "mention_schema_sha256": _sha256(schema_path),
            # The request's counterpart, so a release manifest can attest both
            # ends of the wire rather than only what comes back.
            "request_schema_sha256": _sha256(request_schema_path(config)),
        },
        "versions": {
            "term_family_map": require(config, "term_family.map.version"),
            "grammar": require(config, "grammar.version"),
            "prompt": require(config, "prompt.version"),
            "model_id": require(config, "llm.model.default"),
            "model_revision": require(config, "llm.model.revision"),
            "gateway_revision": require(config, "llm.gateway.revision"),
            # **The serving image, which the contract *can* name.** Unlike the
            # gateway image, the serving image does not carry this file, so
            # naming its digest here is not a hash of a layer containing itself.
            # It is the one image identity that can be an expectation rather
            # than a record.
            "serving_image_digest": require(config, "llm.serving.image_digest"),
            # From the schema's own `$id`, so the contract cannot name a schema
            # other than the one it was compiled against.
            "output_schema": schema["$id"],
            "request_schema": json.loads(
                request_schema_path(config).read_text())["$id"],
            "output_budget_policy": require(config, "llm.output.budget_policy.version"),
        },
        # **The instructions, compiled rather than left in the workbook.**
        # They were authored under `prompt.*` and emitted nowhere, so the
        # contract named a prompt version (`qwen_extractor_v5`) that described
        # text the gateway had no way to send. The model was handed the request
        # document and nothing else — no task, no rules, no schema — which is
        # not a weaker prompt than intended, it is no prompt at all.
        #
        # The few-shots travel too. Their selection policy is a *runtime*
        # decision (`retrieval_policy`) and is left where it is; what belongs
        # here is the material that decision chooses from, so a release manifest
        # attests the same bytes the model was shown.
        "prompt": {
            "system_role": require(config, "prompt.system.role"),
            "system_rules": require(config, "prompt.system.rules"),
            "aboutness_example": require(config, "prompt.output.aboutness_example"),
            "fewshot_validation": require(config, "prompt.fewshot.validation"),
            "fewshots": {
                key.removeprefix("prompt.fewshot.").removesuffix(".output_json"): value
                for key, value in sorted(config.items())
                if key.startswith("prompt.fewshot.")
                and key.endswith(".output_json") and value
            },
        },
        "output_contract": {
            # Schema order, not workbook order: the schema is the wire contract
            # and a reader comparing the two should see the same sequence.
            "families": list(mention["family_hypothesis"]["enum"]),
            "mention_roles": list(mention["mention_role"]["enum"]),
            # **What the model may emit, which under v2 is nothing.** The
            # relation vocabulary itself is not gone — it is
            # `ontology_compiler.model_predicates`, sourced from the grammar
            # sheet, where it has always been. An extraction contract that
            # listed predicates the schema cannot carry would be describing a
            # different schema.
            "predicates": list(
                schema["$defs"]["relation_hypothesis"]["properties"]["predicate"]["enum"]
            ) if "relation_hypothesis" in schema["$defs"] else [],
            "abstain_reasons": _abstain_reasons(schema),
            "max_items_wire": int(require(config, "llm.batch.max_items")),
            "default_batch_items": int(require(config, "llm.batch.default_items")),
            "calibrated_max_items": int(require(config, "llm.batch.calibrated_max_items")),
            "max_mentions_per_item": int(require(config, "llm.max_mentions_per_item")),
            "max_output_tokens": int(require(config, "llm.max_output_tokens")),
            "uncalibrated_item_reserve_tokens":
                int(require(config, "llm.output.uncalibrated_item_reserve_tokens")),
            # **Demanded by `validate` since it was written, and never emitted.**
            # `output_budget_report.py` could not read it and hard-coded 512 in
            # two places instead — one of them an `int(...) and 512` that reads
            # like a derivation and is a constant. A limit the contract requires
            # and does not publish is a limit every reader has to copy.
            "envelope_token_reserve":
                int(require(config, "llm.output.envelope_token_reserve")),
            "predictor_quantile": float(require(config, "llm.output.predictor.quantile")),
            "packing_headroom": float(require(config, "llm.output.packing_headroom")),
            "packing_formula": require(config, "llm.batch.output_budget_formula"),
            "tokenizer_manifest_sha256":
                require(config, "llm.output.tokenizer_manifest_sha256"),
            "stress_fixture_version": require(config, "llm.output.stress_fixture.version"),
            "output_budget_gate": require(config, "release.gate.output_budget"),
        },
        "ontology_compiler": {
            "concept_kind_authority": "live_pg_constraint",
            "concept_kind_target": "ontology.concept_revisions.concept_kind",
            # Sorted: a map has no meaningful order and sorting makes a diff
            # between two compilations readable.
            "family_mappings": {
                key[len("term_family.map."):]: value
                for key, value in sorted(config.items())
                if key.startswith("term_family.map.") and key != "term_family.map.version"
            },
            "hub_canonical": {
                key[len("ontology.hub.canonical."):]: value
                for key, value in config.items()
                if key.startswith("ontology.hub.canonical.")
            },
            "hub_redirects": {
                key[len("ontology.hub.alias."):]: value
                for key, value in config.items()
                if key.startswith("ontology.hub.alias.")
                and not key.endswith((".normalization", ".prompt_policy"))
            },
            "model_predicates": model_predicates(sheets),
        },
        "runtime_requirements": {
            "consumers": _split(require(config, "runtime.contract.required_consumers")),
            "jobs": _split(require(config, "job.pipeline"), ">"),
            "required_storage_objects": _split(require(config, "data.table.minimum")),
            # **Both names, because one of them is not this database's.**
            # The workbook authors these as `private.*`; `private` here is a real
            # schema holding the push secret and the collaborator list. Until
            # now the mapping lived only in `resolve_storage_object`, whose one
            # non-test caller was the gate below — so any other consumer of
            # `required_storage_objects` would read a name that resolves to the
            # wrong schema, which is the same shape of defect as a crosswalk
            # applied silently. Emitted so a repository, a migration checker or
            # a later deploy gate reads the resolved name instead of deriving it
            # again, and so a reviewer can see what was resolved.
            "storage_crosswalk": {
                "schema_map": dict(sorted(STORAGE_SCHEMA_CROSSWALK.items())),
                "objects": [
                    {"declared_name": declared,
                     "production_name": resolve_storage_object(declared)}
                    for declared in _split(require(config, "data.table.minimum"))
                ],
            },
            "initial_mode": require(config, "deployment.initial_mode"),
            "qwen_overlay": require(config, "deployment.feature.semantic_qwen_overlay"),
            "deployment_gates": _split(require(config, "validation.deploy_scope"), "+"),
            "gate_report_schema_version": require(config, "deployment.report.schema_version"),
            "report_test_result_policy": require(config, "deployment.report.test_result_policy"),
            "release_attestation_fields":
                _split(require(config, "deployment.attestation.required_fields")),
            "retrieval_policy": {
                "max_prompt_examples": int(require(config, "runtime_pack.max_examples")),
                "term_cluster_cap": require(config, "runtime_pack.term_cluster_cap"),
                "prompt_registry_mix": require(config, "prompt.fewshot.minimum_active_mix"),
            },
        },
    }


def canonical(contract: dict[str, Any]) -> str:
    """The form the deploy validator compares: everything but `generated_at`."""
    without_timestamp = {k: v for k, v in contract.items() if k != "generated_at"}
    return json.dumps(without_timestamp, indent=2, ensure_ascii=False, sort_keys=False)


#: **`private` in the contract is `semantic_private` in this database.**
#:
#: The contract names its stores `private.review_items` and so on. `private` is a
#: real schema here and holds `push_config` and `collaborators` — the schema
#: nothing is granted on, where the push secret lives — so the contract's name and
#: the production name genuinely differ. It is the same class of crosswalk as
#: `hub:game` meaning `hub:games_play`, and it is written down for the same
#: reason: a compiler that resolved it quietly would be one nobody could audit,
#: and the failure it would hide is a semantic table created next to a secret.
STORAGE_SCHEMA_CROSSWALK = {"private": "semantic_private"}


def resolve_storage_object(declared: str) -> str:
    """Map a contract storage name onto the schema this database actually uses."""
    schema, _, table = declared.partition(".")
    if not table:
        raise ContractError(f"storage object {declared!r} names no schema")
    return f"{STORAGE_SCHEMA_CROSSWALK.get(schema, schema)}.{table}"


#: **Four modes, not a boolean.** The old value was prose
#: (`disabled_until_all_deploy_gates_pass`) tested with `"disabled" in ...`, and
#: that scheme was circular as well as fragile: the overlay was disabled until
#: every gate passed, while `candidate_attestation`, deployed output measurement
#: and `staging_e2e` all require it to have run. It also failed open — a mode
#: named `evaluation` contains no "disabled" and would have read as *enabled*,
#: sending `extract_mentions` straight into `NotImplementedError`.
MODEL_LANE_MODES = ("off", "evaluation", "shadow", "active")

#: Modes in which the model is actually called, so a handler must exist for
#: every registered job.
MODES_CALLING_MODEL = ("evaluation", "shadow", "active")

#: Modes in which rows attributable to a person may be written.
MODES_WRITING_CANDIDATES = ("shadow", "active")


def model_lane_mode(contract: dict[str, Any]) -> str:
    """The declared mode, refused unless it is one of the four.

    Read from the artifact rather than a constant, because the workbook is where
    the decision is authored. **An unrecognised value raises** rather than
    defaulting: the previous substring test treated anything it did not
    understand as *on*, which is the wrong direction for a switch that governs
    whether a model may run against somebody's library.
    """
    mode = str(contract["runtime_requirements"].get("qwen_overlay", "")).strip()
    if mode not in MODEL_LANE_MODES:
        raise ValueError(
            f"model lane mode {mode!r} is not one of {MODEL_LANE_MODES}")
    return mode


def overlay_disabled(contract: dict[str, Any]) -> bool:
    """Kept as the narrow question `pending_jobs` asks: may the model run?"""
    return model_lane_mode(contract) not in MODES_CALLING_MODEL


def pending_jobs(contract: dict[str, Any], live: dict[str, Any]) -> list[str]:
    jobs = set(live.get("job_type") or [])
    if not jobs:
        return []
    missing = [j for j in contract["runtime_requirements"]["jobs"] if j not in jobs]
    if not missing:
        return []
    return ["the live worker_jobs allowlist does not yet permit: " + ", ".join(missing)]


def check_database(contract: dict[str, Any], live: dict[str, Any],
                   pending: list[str] | None = None) -> list[str]:
    """The live-database gate, against values a caller supplies.

    **It takes the live enums as an argument instead of opening a connection**,
    and that is a decision rather than a shortcut. The credential that can read
    `pg_constraint` is not one this repository holds, and a compiler that
    silently falls back to a checked-in snapshot would make
    `concept_kind_authority: live_pg_constraint` a claim rather than a fact.
    Whatever holds the credential reads the constraint and passes it in; this
    checks the contract against it and says which side is wrong.

    The caller must select **only a positive, single-column `ANY(ARRAY[…])`
    allowlist** on the target column. Three constraints in this schema mention
    `concept_kind`, and `booked_activity_candidates_target_binding_v03_check` is
    a compound condition that would yield a wrong, shorter list.
    """
    problems: list[str] = []
    if pending is None:
        pending = []

    # **A hand-written array satisfies this gate exactly as well as a real
    # reading, and looks identical afterwards.** So the snapshot has to say
    # which database it came from and which constraint rows it was parsed out
    # of. `tools/read_live_catalog.py --emit-sql` produces it; anything else has
    # to produce the same evidence, which is the point.
    provenance = live.get("provenance") or {}
    for field in ("database_fingerprint_sha256", "constraint_oids",
                  "database", "server_version"):
        if not provenance.get(field):
            problems.append(
                f"the live snapshot carries no {field}; it was not read from a "
                f"database by tools/read_live_catalog.py"
            )

    kinds = set(live.get("concept_kind") or [])
    if not kinds:
        problems.append("no live concept_kind allowlist supplied")

    for family, mapping in contract["ontology_compiler"]["family_mappings"].items():
        parts = dict(
            piece.split("=", 1) for piece in mapping.split(";") if "=" in piece
        )
        kind = parts.get("concept_kind")
        if parts.get("storage") != "concepts":
            continue
        if kind and kind not in kinds:
            problems.append(
                f"family {family!r} compiles to concept_kind={kind!r}, "
                f"which the live constraint does not permit"
            )

    # **Jobs are the one check that measures a future state, so it is reported
    # separately rather than folded in.** Eight of the nine pipeline jobs belong
    # to the overlay the contract itself declares
    # `disabled_until_all_deploy_gates_pass`, and registering a `job_type` before
    # its handler ships is worse than not registering it — a claimed job is
    # claimed, found to have no handler, and marked `dead` without retry.
    #
    # So while the overlay is off, their absence is *pending* and not a fault;
    # the moment it is turned on, the same absence is a release blocker. A gate
    # that stays red on a known-unbuilt thing is one people stop reading, and a
    # gate that goes quiet when the thing gets built is worse.
    pending.extend(pending_jobs(contract, live))
    if not overlay_disabled(contract):
        problems.extend(pending_jobs(contract, live))

    hubs = set(live.get("hubs") or [])
    if hubs:
        for role, hub in contract["ontology_compiler"]["hub_canonical"].items():
            if hub not in hubs:
                problems.append(f"canonical hub {hub!r} for role {role!r} is not in the ontology")
        for legacy in contract["ontology_compiler"]["hub_redirects"]:
            if legacy in hubs:
                problems.append(f"legacy hub {legacy!r} exists in the ontology and must be merged")

    # The `storage_integration` gate's first half: the sixteen stores exist,
    # under the names this database uses rather than the ones the contract does.
    tables = set(live.get("tables") or [])
    if tables:
        for declared in contract["runtime_requirements"]["required_storage_objects"]:
            actual = resolve_storage_object(declared)
            if actual not in tables:
                note = "" if actual == declared else f" (declared as {declared})"
                problems.append(f"required storage object {actual} does not exist{note}")

    # **A check constraint restating a closed vocabulary is a second copy of it**,
    # and this repository's recurring defect is one fact in two places. The copy
    # in `semantic_private.provisional_entities.family` is permitted because the
    # database has no other mechanism — and it is made safe here, by reading it
    # back and refusing to agree that the contract compiles if the two have
    # drifted. The contract is the authority; the constraint is its enforcement.
    # The same argument as the families below, one table over. `0207` restates
    # the contract's fifteen mention roles as a check constraint because the
    # database has no other mechanism, and the copy is made safe by being read
    # back. Two vocabularies share that column — the legacy resolver's field
    # names and the contract's roles — and this is what keeps the second half
    # honest.
    stored_roles = set(live.get("mention_role") or [])
    if stored_roles:
        declared_roles = set(contract["output_contract"]["mention_roles"])
        for extra in sorted(stored_roles - declared_roles):
            problems.append(
                f"observation_mentions permits mention_role {extra!r}, which the "
                f"contract does not declare"
            )
        for absent in sorted(declared_roles - stored_roles):
            problems.append(
                f"the contract declares mention_role {absent!r} and "
                f"observation_mentions refuses it"
            )

    stored_families = set(live.get("provisional_family") or [])
    if stored_families:
        declared_families = set(contract["ontology_compiler"]["family_mappings"])
        for extra in sorted(stored_families - declared_families):
            problems.append(
                f"provisional_entities.family permits {extra!r}, which no family mapping declares"
            )
        for absent in sorted(declared_families - stored_families):
            problems.append(
                f"family {absent!r} is declared but provisional_entities.family refuses it"
            )
    return problems


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--emit", action="store_true", help="rewrite the compiled contract")
    parser.add_argument("--check", action="store_true",
                        help="validate and compare against the checked-in contract")
    parser.add_argument("--check-database", action="store_true",
                        help="validate the contract against live enums")
    parser.add_argument("--live-enums",
                        help="JSON with concept_kind[], job_type[] and hubs[] read from the "
                             "live database by whatever holds the credential")
    arguments = parser.parse_args()

    # **The database gate reads the shipped contract and never the workbook**, and
    # that separation is worth keeping deliberately. The question it asks is
    # whether the artifact production consumes agrees with the schema that is
    # actually built — the authoring spreadsheet has no standing in it. Loading
    # the workbook first made the gate need `openpyxl`, which meant it could not
    # run inside `tools/replay_contracts.sh`, which is the one place holding a
    # database built from nothing but the migrations.
    if arguments.check_database:
        if not arguments.live_enums:
            print("--check-database requires --live-enums", file=sys.stderr)
            return 2
        existing = json.loads(CONTRACT.read_text())
        pending: list[str] = []
        problems = check_database(
            existing,
            json.loads(pathlib.Path(arguments.live_enums).read_text()),
            pending,
        )
        for note in pending:
            print(f"  live database (pending, overlay off): {note}", file=sys.stderr)
        for problem in problems:
            print(f"  live database: {problem}", file=sys.stderr)
        print(f"live database: {len(problems)} problem(s), {len(pending)} pending",
              file=sys.stderr)
        return 1 if problems else 0

    sheets = load_workbook()
    try:
        schema_path = output_schema_path(config_of(sheets))
    except ContractError as refusal:
        print(f"contract refused: {refusal}", file=sys.stderr)
        return 1
    schema = json.loads(schema_path.read_text())
    try:
        request_schema = json.loads(request_schema_path(config_of(sheets)).read_text())
    except ContractError as refusal:
        print(f"contract refused: {refusal}", file=sys.stderr)
        return 1

    try:
        validate(sheets, schema, request_schema)
    except ContractError as refusal:
        print(f"contract refused: {refusal}", file=sys.stderr)
        return 1
    print("artifact parity: workbook, schema and grammar agree", file=sys.stderr)

    from datetime import datetime, timezone
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%S.") + \
        f"{datetime.now(timezone.utc).microsecond // 1000:03d}Z"
    contract = compile_contract(sheets, schema, schema_path, stamp)

    if arguments.emit:
        CONTRACT.write_text(json.dumps(contract, indent=2, ensure_ascii=False) + "\n")
        print(f"wrote {CONTRACT.relative_to(REPOSITORY)}", file=sys.stderr)
        return 0

    if arguments.check:
        existing = json.loads(CONTRACT.read_text())
        if canonical(existing) != canonical(contract):
            print("the checked-in contract is not what these artifacts compile to",
                  file=sys.stderr)
            return 1
        print("compiled contract matches the checked-in artifact", file=sys.stderr)
        return 0

    json.dump(contract, sys.stdout, indent=2, ensure_ascii=False)
    print()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
